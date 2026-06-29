import os
import io
import base64
import json
import tempfile

from flask import Flask, render_template, request, jsonify, send_file
import fitz  # PyMuPDF
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

HEADERS = [
    'No.',
    'Last Name',
    'First Name',
    'Company',
    'Department',
    'Job Role',
    '1. 오늘 세션에서 다룬 내용의 깊이와 범위는 어떠셨습니까? ',
    '1-1. 세미나 진행 관련, 제안사항이나 의견을 공유해 주실 사항이 있다면 자유롭게 작성 부탁드립니다.',
    '2. 금일 세미나에서 소개� Red Hat 솔루션에 대해 추가 안내 받기를 원하십니까?',
    '3. 원하시는 상담 유형을 선택해 주세요.\n',
    '4. Red Hat 세션에서 듣고 싶으신 다른 주제가 있으시다면 안늕 추가 안내 가능니다. \n(중복 선택해 가능)',
    '정보 제공 동흘',
]


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_pdf():
    if 'pdf' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400

    file = request.files['pdf']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'PDF 파일만 업로드 가능합니다.'}), 400

    pdf_bytes = file.read()
    doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    pages = []

    for i, page in enumerate(doc):
        mat = fitz.Matrix(2.5, 2.5)
        pix = page.get_pixmap(matrix=mat)
        img_data = pix.tobytes('png')
        img_b64 = base64.b64encode(img_data).decode('utf-8')

        pages.append({
            'index': i,
            'image': f'data:image/png;base64,{img_b64}'
        })

    doc.close()
    return jsonify({'pages': pages, 'total': len(pages)})


@app.route('/export', methods=['POST'])
def export_excel():
    data = request.get_json()
    respondents = data.get('respondents', [])

    wb = Workbook()
    ws = wb.active
    ws.title = '설문 (응답낔욙)'

    # 헤�T 스타일
    header_fill = PatternFill('solid', start_color='CC0000', end_color='CC0000')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 40

    # 데이터 행
    data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for r in respondents:
        q4_val = ', '.join(r.get('q4', [])) if isinstance(r.get('q4'), list) else r.get('q4', '')
        row = [
            r.get('no', ''),
            r.get('lastName', ''),
            r.get('firstName', ''),
            r.get('company', ''),
            r.get('department', ''),
            r.get('jobRole', ''),
            r.get('q1', ''),
            r.get('q1_1', ''),
            r.get('q2', ''),
            r.get('q3', ''),
            q4_val,
            r.get('consent', ''),
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = data_align
            cell.border = border

    # 쫬럼 너비 조정
    col_widths = [6, 10, 12, 18, 22, 20, 20, 30, 14, 28, 40, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='redhat_survey.xlsx'
    )


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
